#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Скрипт для отладки шаблона
"""

import openpyxl

# Загружаем шаблон
wb = openpyxl.load_workbook('templates/employee_template v4.0 lightweight.xlsx')
print('Листы:', wb.sheetnames)

# Проверяем лист "Расчёт"
if 'Расчёт' in wb.sheetnames:
    ws = wb['Расчёт']
    print('Ячейка B4:', ws['B4'].value)
    print('Ячейка B10:', ws['B10'].value)
    print('Ячейка C4:', ws['C4'].value)
    print('Ячейка C10:', ws['C10'].value)
else:
    print('Лист "Расчёт" не найден')

# Проверяем лист "rules"
if 'rules' in wb.sheetnames:
    rules_ws = wb['rules']
    print('\nПравила:')
    for row in range(2, 10):
        target = rules_ws.cell(row=row, column=1).value
        source = rules_ws.cell(row=row, column=2).value
        rule_type = rules_ws.cell(row=row, column=3).value
        if target and source and rule_type:
            print(f'Строка {row}: {target} -> {source} ({rule_type})')

wb.close() 