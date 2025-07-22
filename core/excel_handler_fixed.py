# -*- coding: utf-8 -*-
"""
Модуль работы с Excel файлами
"""

import logging
import shutil
from pathlib import Path
from datetime import datetime, date
from typing import List, Optional, Dict, Tuple, Any
import re

import openpyxl
from openpyxl.styles import Border, Side

from models import Employee, VacationInfo, VacationPeriod, VacationStatus
from config import Config
from core.performance_tracker import PerformanceTracker, FilePerformanceStats
from core.directory_manager import DirectoryManager
from core.data_mapper import DataMapper


class ExcelHandler:
    """Класс для работы с Excel файлами"""

    def __init__(self, config):
        self.config = config
        self.logger = logging.getLogger(__name__)
        self._cached_rules = {}
        self._cached_templates = {}
        self._cached_workbooks = {}
        self._cached_cell_addresses = {}  # Кэш для парсинга адресов ячеек
        self.performance_tracker = PerformanceTracker()
        self.directory_manager = DirectoryManager(config)
        self.data_mapper = DataMapper()
    
    def _get_cached_rules(self, template_path: str) -> Dict[str, Dict[str, str]]:
        """Получает rules из кэша или загружает их"""
        if template_path not in self._cached_rules:
            self._cached_rules[template_path] = self._load_filling_rules(template_path)
        return self._cached_rules[template_path]
    
    def _get_cached_template_workbook(self, template_path: str) -> openpyxl.Workbook:
        """Получает шаблон из кэша или загружает его"""
        if template_path not in self._cached_workbooks:
            self._cached_workbooks[template_path] = openpyxl.load_workbook(
                template_path,
                data_only=False,
                read_only=True,
                keep_links=False
            )
        return self._cached_workbooks[template_path]
    
    def create_employee_file(self, employee: Dict[str, str], output_path: str) -> bool:
        """Создает файл сотрудника на основе шаблона с rules"""
        file_stats = self.performance_tracker.start_file(employee['ФИО работника'])
        
        try:
            template_path = Path(self.config.employee_template)
            
            if not template_path.exists():
                self.logger.error(f"Шаблон не найден: {template_path}")
                file_stats.finish(False, f"Шаблон не найден: {template_path}")
                raise FileNotFoundError(f"Шаблон сотрудника не найден: {template_path}")
            
            self.directory_manager.ensure_directory_exists(Path(output_path).parent)
            
            # Используем кэшированный шаблон для копирования
            cached_template = self._get_cached_template_workbook(str(template_path))
            shutil.copy2(template_path, output_path)
            
            # Получаем кэшированные rules
            rules = self._get_cached_rules(str(template_path))
            
            # Подготавливаем данные сотрудника
            data_dict = {}
            for field_name in rules.get('value', {}).values():
                value = employee.get(field_name, '')
                if value is None:
                    value = ''
                data_dict[field_name] = value
            
            # Добавляем даты отпусков из входного файла
            if 'vacation_dates' in employee:
                data_dict['vacation_dates'] = employee['vacation_dates']
            
            # Загружаем файл для редактирования
            workbook = openpyxl.load_workbook(
                output_path,
                data_only=False,
                read_only=False,
                keep_links=False
            )
            
            # Применяем правила заполнения
            self._apply_rules_to_template(workbook, rules, data_dict)
            
            workbook.save(output_path)
            workbook.close()
            
            file_stats.finish(True)
            return True
            
        except Exception as e:
            self.logger.error(f"Ошибка при создании файла {employee['ФИО работника']}: {e}")
            file_stats.finish(False, str(e))
            return False

    def _apply_rules_to_template(self, workbook, rules: Dict[str, Dict[str, str]], data_dict: Dict[str, Any]):
        """Применяет правила заполнения к шаблону"""
        
        for rule_type, rule_items in rules.items():
            if rule_type == 'value':
                for cell_address, field_name in rule_items.items():
                    value = data_dict.get(field_name, '')
                    
                    try:
                        is_formula, clean_address, sheet_name = self._parse_cell_address(cell_address)
                        self._fill_cell_or_range(workbook, sheet_name, clean_address, value)
                        
                    except Exception as e:
                        self.logger.error(f"Ошибка при заполнении {cell_address}: {e}")
        
        # Заполняем даты отпусков из входного файла
        if 'vacation_dates' in data_dict and data_dict['vacation_dates']:
            self._fill_vacation_dates(workbook, data_dict['vacation_dates'])

    def _parse_cell_address(self, address: str) -> tuple:
        """Парсит адрес ячейки, возвращает (is_formula, clean_address, sheet_name)"""
        # Проверяем кэш
        if address in self._cached_cell_addresses:
            return self._cached_cell_addresses[address]
        
        is_formula = address.startswith('=')
        
        if is_formula:
            # Убираем знак равенства и извлекаем адрес
            formula = address[1:]
            
            # Простой парсинг для формул вида 'Лист'!A1
            if '!' in formula:
                sheet_part, cell_part = formula.split('!', 1)
                sheet_name = sheet_part.strip("'\"")
                clean_address = cell_part.strip()
            else:
                sheet_name = None
                clean_address = formula.strip()
        else:
            # Обычный адрес - может быть именованный диапазон
            sheet_name = None
            clean_address = address.strip()
        
        result = (is_formula, clean_address, sheet_name)
        self._cached_cell_addresses[address] = result
        return result

    def _convert_value_type(self, value: Any) -> Any:
        """Преобразует значение к правильному типу данных для Excel"""
        if value is None or value == '':
            return ''
        
        if isinstance(value, (int, float)):
            # Сохраняем целые числа как int, float как float
            return value
        
        str_value = str(value).strip()
        
        if not str_value:
            return ''
        
        # Проверяем является ли параметр числом
        try:
            # Убираем пробелы и неразрывные пробелы
            clean_value = str_value.replace(' ', '').replace('\xa0', '')
            # Заменяем запятую на точку
            clean_value = clean_value.replace(',', '.')
            # Пытаемся преобразовать в число
            float_val = float(clean_value)
            # Если это целое число, возвращаем int
            if float_val.is_integer():
                return int(float_val)
            else:
                return float_val
        except (ValueError, TypeError):
            pass
        
        return str_value
    
    def _fill_cell_fast(self, worksheet, address: str, value: Any) -> None:
        """Быстрое заполнение ячейки без сложных проверок"""
        try:
            cell = worksheet[address]
            cell.value = value
        except Exception as e:
            self.logger.error(f"Ошибка при быстром заполнении {address}: {e}")
    
    def _fill_cell_or_range(self, workbook, sheet_name: str, address: str, value: Any) -> None:
        """Заполняет ячейку или диапазон значением с правильным типом данных"""
        converted_value = self._convert_value_type(value)
        
        try:
            if sheet_name:
                if sheet_name in workbook.sheetnames:
                    worksheet = workbook[sheet_name]
                else:
                    worksheet = workbook.worksheets[0]
            else:
                worksheet = workbook.worksheets[0]
            
            # Проверяем тип адреса
            if re.match(r'^[A-Z]+[0-9]+$', address):
                # Стандартный адрес ячейки
                cell = worksheet[address]
                cell.value = converted_value
                # Принудительно устанавливаем тип данных
                if isinstance(converted_value, (int, float)):
                    cell.data_type = 'n'
                    cell.number_format = '0' if isinstance(converted_value, int) else '0.00'
                elif isinstance(converted_value, str) and converted_value != '':
                    cell.data_type = 's'
                    
            elif ':' in address:
                # Диапазон - заполняем первую ячейку
                start_cell = address.split(':')[0]
                if re.match(r'^[A-Z]+[0-9]+$', start_cell):
                    cell = worksheet[start_cell]
                    cell.value = converted_value
                    if isinstance(converted_value, (int, float)):
                        cell.data_type = 'n'
                        cell.number_format = '0' if isinstance(converted_value, int) else '0.00'
                    elif isinstance(converted_value, str) and converted_value != '':
                        cell.data_type = 's'
            else:
                # Именованный диапазон
                if address in workbook.defined_names:
                    defn = workbook.defined_names[address]
                    if defn.attr_text:
                        range_text = defn.attr_text
                        
                        if '!' in range_text:
                            sheet_part, cell_part = range_text.split('!', 1)
                            cell_part = cell_part.replace('$', '')
                            if re.match(r'^[A-Z]+[0-9]+$', cell_part):
                                cell = worksheet[cell_part]
                                cell.value = converted_value
                                if isinstance(converted_value, (int, float)):
                                    cell.data_type = 'n'
                                    cell.number_format = '0' if isinstance(converted_value, int) else '0.00'
                                elif isinstance(converted_value, str) and converted_value != '':
                                    cell.data_type = 's'
                else:
                    # Попытка заполнить как обычную ячейку
                    cell = worksheet[address]
                    cell.value = converted_value
                    if isinstance(converted_value, (int, float)):
                        cell.data_type = 'n'
                        cell.number_format = '0' if isinstance(converted_value, int) else '0.00'
                    elif isinstance(converted_value, str) and converted_value != '':
                        cell.data_type = 's'
                    
        except Exception as e:
            self.logger.error(f"Ошибка при заполнении {address}: {e}")
            raise e
    
    def _is_float(self, value: str) -> bool:
        """Проверяет, является ли строка числом с плавающей точкой"""
        try:
            float(value)
            return True
        except (ValueError, TypeError):
            return False
            
    def _fill_vacation_dates(self, workbook, vacation_dates: List[Dict[str, Any]]):
        """Заполняет даты отпусков в файле сотрудника"""
        # ИСПРАВЛЕНО: В правильном входном файле нет дат отпусков
        # Даты отпусков будут заполняться вручную в формах сотрудников
        if not vacation_dates:
            return
            
        try:
            worksheet = workbook.worksheets[0]  # Основной лист
            
            # Заполняем даты в строках 15-29 только если они есть
            for i, vacation in enumerate(vacation_dates[:15]):  # Максимум 15 периодов
                row = 15 + i
                
                start_date = vacation.get('start_date')
                end_date = vacation.get('end_date')
                days = vacation.get('days')
                
                if start_date:
                    worksheet[f"C{row}"] = start_date
                if end_date:
                    worksheet[f"D{row}"] = end_date
                if days:
                    worksheet[f"E{row}"] = days
                    
        except Exception as e:
            self.logger.error(f"Ошибка при заполнении дат отпусков: {e}")
        
    def clear_cache(self) -> None:
        """Очищает кэш для освобождения памяти"""
        for workbook in self._cached_workbooks.values():
            try:
                workbook.close()
            except Exception:
                pass
        
        self._cached_workbooks.clear()
        self._cached_rules.clear()
        self._cached_templates.clear()
        self._cached_cell_addresses.clear()

    def _load_filling_rules(self, template_path: str) -> Dict[str, Dict[str, str]]:
        """Загружает правила заполнения из листа 'rules'"""
        rules = {'value': {}, 'header': {}, 'read': {}}
        
        workbook = openpyxl.load_workbook(template_path, data_only=False)
        
        if 'rules' not in workbook.sheetnames:
            workbook.close()
            raise ValueError(f"Лист 'rules' не найден в шаблоне {template_path}")
        
        rules_sheet = workbook['rules']
        
        for row in range(2, rules_sheet.max_row + 1):
            target_cell = rules_sheet.cell(row=row, column=1)  # Столбец A - КУДА
            source_cell = rules_sheet.cell(row=row, column=2)  # Столбец B - ЧТО
            type_cell = rules_sheet.cell(row=row, column=3)    # Столбец C - ТИП
            
            target_address = target_cell.value if target_cell.data_type == 'f' else target_cell.value
            source_field = source_cell.value
            rule_type = type_cell.value
            
            if target_address and source_field and rule_type:
                target_address = str(target_address).strip()
                source_field = str(source_field).strip()
                rule_type = str(rule_type).strip().lower()
                
                if rule_type in ['value', 'header', 'read']:
                    rules[rule_type][target_address] = source_field
        
        workbook.close()
        
        if not any(rules.values()):
            raise ValueError(f"Лист 'rules' пуст или не содержит корректных правил в {template_path}")
        
        return rules

     def create_block_report(self, block_name: str, vacation_infos: List[VacationInfo], output_path: str) -> bool:
        """Создает отчет по блоку с использованием rules"""
        template_path = Path(self.config.block_report_template)
        if not template_path.exists():
            raise FileNotFoundError(f"Шаблон отчета не найден: {template_path}")
        self.directory_manager.ensure_directory_exists(Path(output_path).parent)
        shutil.copy2(template_path, output_path)
        rules = self._load_filling_rules(str(template_path))
        workbook = openpyxl.load_workbook(output_path)
        self._fill_report_with_rules(workbook, block_name, vacation_infos, rules)
        workbook.save(output_path)
        workbook.close()
        return True

    def _fill_report_with_rules(self, workbook, block_name: str, vacation_infos: List[VacationInfo], rules: Dict[str, Dict[str, str]]):
        """Заполняет отчет используя rules"""
        # Используем DataMapper для динамического маппинга заголовка
        report_data = self.data_mapper.map_report_header_data(block_name, vacation_infos)
        
        # Применяем rules
        self._apply_rules_to_template(workbook, rules, report_data)
        
        # Заполняем таблицы данных
        self._fill_employee_tables(workbook, vacation_infos, rules)
        
        # Заполняем календарь
        if 'Report' in workbook.sheetnames:
            self._fill_calendar_matrix(workbook['Report'], vacation_infos)

    def _fill_employee_tables(self, workbook, vacation_infos: List[VacationInfo], rules: Dict[str, Dict[str, str]]):
        """Заполняет таблицы сотрудников на Report и Print листах"""
        if 'Report' in workbook.sheetnames:
            self._fill_table_by_prefix(workbook['Report'], vacation_infos, rules, 'report_', self._get_report_row_data_dynamic)
            self._apply_borders_to_report_table(workbook['Report'], len(vacation_infos), rules)
        
        if 'Print' in workbook.sheetnames:
            normalized_data = self._normalize_vacation_data(vacation_infos)
            self._fill_table_by_prefix(workbook['Print'], normalized_data, rules, 'print_', self._get_print_row_data_dynamic)
            self._apply_borders_to_table(workbook['Print'], len(normalized_data))

    def _fill_table_by_prefix(self, worksheet, data_list: List, rules: Dict[str, Dict[str, str]], prefix: str, row_data_func):
        """Универсальная функция заполнения таблицы по префиксу"""
        header_rules = rules.get('header', {})
        
        # Собираем mapping: имя поля -> столбец
        column_mapping = {}
        for cell_address, field_name in header_rules.items():
            if field_name.startswith(prefix):
                if cell_address.startswith('=') and '!' in cell_address:
                    sheet_and_cell = cell_address[1:]
                    if '!' in sheet_and_cell:
                        sheet_part, cell_part = sheet_and_cell.split('!', 1)
                        col_match = re.search(r'([A-Z]+)', cell_part)
                        row_match = re.search(r'(\d+)', cell_part)
                        if col_match and row_match:
                            column_mapping[field_name] = (col_match.group(1), int(row_match.group(1)))
                else:
                    col_match = re.search(r'([A-Z]+)', cell_address)
                    row_match = re.search(r'(\d+)', cell_address)
                    if col_match and row_match:
                        column_mapping[field_name] = (col_match.group(1), int(row_match.group(1)))
        
        if not column_mapping:
            return
        
        # Заполняем данные
        for i, data_item in enumerate(data_list):
            row_data = row_data_func(data_item, i)
            
            for key, value in row_data.items():
                if key in column_mapping:
                    col, header_row = column_mapping[key]
                    data_row = header_row + 1 + i
                    cell_address = f"{col}{data_row}"
                    
                    # Преобразуем значение к правильному типу
                    converted_value = self._convert_value_type(value)
                    worksheet[cell_address] = converted_value

    def _get_report_row_data_dynamic(self, vacation_info: VacationInfo, index: int) -> Dict[str, Any]:
        """Динамически получает данные строки для Report листа используя DataMapper"""
        return self.data_mapper.map_vacation_info_to_rules(vacation_info, index, 'report_')
    
    def _get_report_row_data(self, vacation_info: VacationInfo, index: int) -> Dict[str, Any]:
        emp = vacation_info.employee
        
        return {
            "report_row_number": index + 1,
            "report_employee_name": emp.get("ФИО работника", ""),
            "report_tab_number": emp.get("Табельный номер", ""),
            "report_position": emp.get("Должность", ""),
            "report_department1": emp.get("Подразделение 1", ""),
            "report_department2": emp.get("Подразделение 2", ""),
            "report_department3": emp.get("Подразделение 3", ""),
            "report_department4": emp.get("Подразделение 4", ""),
            "report_status": vacation_info.get_status_text(),
            "report_total_days": sum(p.days for p in vacation_info.periods) if vacation_info.periods else "",
            "report_periods_count": len(vacation_info.periods) if vacation_info.periods else "",
        }

    def _get_print_row_data_dynamic(self, data, index: int) -> Dict[str, Any]:
        """Динамически получает данные строки для Print листа используя DataMapper"""
        return self.data_mapper.map_period_data_to_rules(data, index, 'print_')
    
    def _get_print_row_data(self, data, index: int) -> Dict[str, Any]:
        """Получает данные строки для Print листа из нормализованных данных"""
        emp = data.get('employee', {})
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        days = data.get('days', 0)
        
        start_date_str = start_date.strftime('%d.%m.%Y') if start_date else ""
        end_date_str = end_date.strftime('%d.%m.%Y') if end_date else ""
        
        return {
            "print_row_number": index + 1,
            "print_tab_number": emp.get("Табельный номер", ""),
            "print_employee_name": emp.get("ФИО работника", ""),
            "print_position": emp.get("Должность", ""),
            "print_start_date": start_date_str,
            "print_end_date": end_date_str,
            "print_duration": str(days) if days > 0 else "",
            "print_signature": "",
            "print_acknowledgment_date": "",
            "print_notes": "",
        }

    def _normalize_vacation_data(self, vacation_infos: List[VacationInfo]) -> List[Dict]:
        """Нормализует данные отпусков - каждый период = отдельная строка"""
        normalized_data = []
        for vacation_info in vacation_infos:
            emp = vacation_info.employee
            if not vacation_info.periods:
                normalized_data.append({
                    'employee': emp, 'period_num': 0, 'start_date': None, 'end_date': None, 'days': 0
                })
            else:
                for period_idx, period in enumerate(vacation_info.periods, 1):
                    normalized_data.append({
                        'employee': emp, 'period_num': period_idx,
                        'start_date': period.start_date, 'end_date': period.end_date, 'days': period.days
                    })
        return normalized_data

    def _apply_borders_to_report_table(self, worksheet, data_count: int, rules: Dict[str, Dict[str, str]]):
        """Применяет границы к таблице Report листа"""
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                           top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Получаем mapping столбцов из rules
        header_rules = rules.get('header', {})
        column_mapping = {}
        
        for cell_address, field_name in header_rules.items():
            if field_name.startswith('report_'):
                if cell_address.startswith('=') and '!' in cell_address:
                    sheet_and_cell = cell_address[1:]
                    if '!' in sheet_and_cell:
                        sheet_part, cell_part = sheet_and_cell.split('!', 1)
                        col_match = re.search(r'([A-Z]+)', cell_part)
                        row_match = re.search(r'(\d+)', cell_part)
                        if col_match and row_match:
                            column_mapping[field_name] = (col_match.group(1), int(row_match.group(1)))
                else:
                    col_match = re.search(r'([A-Z]+)', cell_address)
                    row_match = re.search(r'(\d+)', cell_address)
                    if col_match and row_match:
                        column_mapping[field_name] = (col_match.group(1), int(row_match.group(1)))
        
        if not column_mapping:
            return
        
        # Находим диапазон столбцов и строк
        min_row = float('inf')
        max_row = 0
        min_col = float('inf')
        max_col = 0
        
        for col_str, header_row in column_mapping.values():
            # Конвертируем буквенное обозначение столбца в номер
            col_num = 0
            for char in col_str:
                col_num = col_num * 26 + (ord(char.upper()) - ord('A') + 1)
            
            min_col = min(min_col, col_num)
            max_col = max(max_col, col_num)
            min_row = min(min_row, header_row + 1)  # +1 потому что данные начинаются со следующей строки
            max_row = max(max_row, header_row + 1 + data_count - 1)  # последняя строка с данными
        
        # Применяем границы ко всем ячейкам в диапазоне
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = thin_border

    def _apply_borders_to_table(self, worksheet, data_count: int):
        """Применяет границы к таблице Print листа"""
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                           top=Side(style='thin'), bottom=Side(style='thin'))
        
        report_structure = self.config.report_structure
        start_row = report_structure.get("employee_data_start_row", 9)
        
        for row in range(start_row, start_row + data_count):
            for col in range(1, 11):
                worksheet.cell(row=row, column=col).border = thin_border

    def _fill_calendar_matrix(self, worksheet, vacation_infos: List[VacationInfo]):
        """Заполняет календарную матрицу"""
        report_structure = self.config.report_structure
        start_col = report_structure.get("calendar_start_col", 12)
        month_row = report_structure.get("calendar_month_row", 7)
        day_row = report_structure.get("calendar_day_row", 8)
        employee_start_row = report_structure.get("employee_data_start_row", 9)
        
        month_names = self.config.month_names
        days_in_months = self.config.days_in_months
        target_year = self.config.target_year
        
        col_offset = 0
        for month_idx, month_name in enumerate(month_names):
            month_col = start_col + col_offset
            worksheet.cell(row=month_row, column=month_col, value=month_name)
            
            days_in_month = days_in_months[month_idx]
            for day in range(1, days_in_month + 1):
                day_col = start_col + col_offset + day - 1
                worksheet.cell(row=day_row, column=day_col, value=day)
            
            col_offset += days_in_month
        
        for emp_idx, vacation_info in enumerate(vacation_infos):
            emp_row = employee_start_row + emp_idx
            for period in vacation_info.periods:
                current_date = period.start_date
                while current_date <= period.end_date:
                    if current_date.year == target_year:
                        day_col = self._get_calendar_column(current_date, start_col)
                        if day_col:
                            worksheet.cell(row=emp_row, column=day_col, value=1)
                    
                    from datetime import timedelta
                    current_date = current_date + timedelta(days=1)
                    if current_date > period.end_date:
                        break

    def _get_calendar_column(self, target_date: date, start_col: int) -> Optional[int]:
        """Вычисляет столбец для даты в календаре"""
        target_year = self.config.target_year
        days_in_months = self.config.days_in_months
        
        if target_date.year != target_year:
            return None
        
        col_offset = sum(days_in_months[:target_date.month - 1]) + target_date.day - 1
        return start_col + col_offset

    def read_block_report_data(self, report_path: str) -> Optional[Dict]:
        """Читает данные из отчета по блоку"""
        try:
            workbook = openpyxl.load_workbook(report_path, data_only=True, read_only=True)
            if 'Report' not in workbook.sheetnames:
                return None
            
            worksheet = workbook['Report']
            
            report_structure = self.config.report_structure
            header_cells = report_structure.get("report_header_cells", {})
            
            block_name = str(self._get_cell_value(worksheet, header_cells.get("block_name", "A3")) or "").strip()
            update_date_raw = str(self._get_cell_value(worksheet, header_cells.get("update_date", "A4")) or "").strip()
            total_employees_raw = str(self._get_cell_value(worksheet, header_cells.get("total_employees", "A5")) or "").strip()
            completed_raw = str(self._get_cell_value(worksheet, header_cells.get("completed", "A6")) or "").strip()
            
            update_date = update_date_raw.replace("Дата обновления:", "").strip() if "Дата обновления:" in update_date_raw else ""
            
            total_employees = 0
            if "Количество сотрудников:" in total_employees_raw:
                try:
                    total_employees = int(total_employees_raw.split(":")[1].strip())
                except (ValueError, IndexError):
                    pass
            
            completed_employees = 0
            percentage = 0
            if "Закончили планирование:" in completed_raw:
                try:
                    parts = completed_raw.split(":")[1].strip().split("(")
                    completed_employees = int(parts[0].strip())
                    if len(parts) > 1:
                        percentage = int(parts[1].replace("%)", "").strip())
                except (ValueError, IndexError):
                    pass
            
            workbook.close()
            
            return {
                'block_name': block_name,
                'total_employees': total_employees,
                'completed_employees': completed_employees,
                'remaining_employees': total_employees - completed_employees,
                'percentage': percentage,
                'update_date': update_date
            }
            
        except Exception as e:
            self.logger.error(f"Ошибка чтения отчета {report_path}: {e}")
            return None

    def read_block_report_data_by_rules(self, report_path: str) -> Optional[Dict]:
        """Читает данные из отчета по блоку используя его rules"""
        try:
            # Загружаем rules отчета по блоку
            block_rules = self._load_filling_rules(report_path)
            
            workbook = openpyxl.load_workbook(report_path, data_only=True, read_only=True)
            if 'Report' not in workbook.sheetnames:
                return None
            
            worksheet = workbook['Report']
            
            # Читаем данные по rules
            data = {}
            
            # Читаем value правила (основные данные отчета)
            value_rules = block_rules.get('value', {})
            for cell_address, field_name in value_rules.items():
                try:
                    is_formula, clean_address, sheet_name = self._parse_cell_address(cell_address)
                    if sheet_name is None:
                        sheet_name = 'Report'
                    
                    if sheet_name in workbook.sheetnames:
                        cell = workbook[sheet_name][clean_address]
                        if isinstance(cell, tuple):
                            value = cell[0].value if cell else None
                        elif hasattr(cell, 'value'):
                            value = cell.value
                        else:
                            value = cell
                        data[field_name] = value
                except Exception as e:
                    self.logger.warning(f"Ошибка чтения {cell_address}: {e}")
            
            # Название блока должно быть в value правилах как block_name
            # Если нет, то используем пустое значение
            if 'block_name' not in data:
                data['block_name'] = ""
            
            workbook.close()
            
            # Подготавливаем результат
            block_name = data.get('block_name', '')
            total_employees = int(data.get('total_employees', 0))
            employees_filled = int(data.get('employees_filled', 0))
            employees_correct = int(data.get('employees_correct', 0))
            update_date = str(data.get('update_date', ''))
            
            # Вычисляем дополнительные поля
            remaining_employees = total_employees - employees_correct
            percentage = round((employees_correct / total_employees * 100) if total_employees > 0 else 0, 0)
            
            return {
                'block_name': block_name,
                'total_employees': total_employees,
                'completed_employees': employees_correct,
                'remaining_employees': remaining_employees,
                'percentage': percentage,
                'update_date': update_date,
                'employees_filled': employees_filled,
                'employees_incorrect': employees_filled - employees_correct,
                'employees_not_filled': total_employees - employees_filled
            }
            
        except Exception as e:
            self.logger.error(f"Ошибка чтения отчета {report_path}: {e}")
            return None

    def create_general_report_from_blocks(self, block_data: List[Dict], output_path: str) -> bool:
        """Создает общий отчет"""
        template_path = Path(self.config.general_report_template)
        if not template_path.exists():
            raise FileNotFoundError(f"Шаблон общего отчета не найден: {template_path}")
        
        self.directory_manager.ensure_directory_exists(Path(output_path).parent)
        shutil.copy2(template_path, output_path)
        
        workbook = openpyxl.load_workbook(output_path)
        worksheet = workbook.active
        
        if worksheet is None:
            return False
        
        # Заполняем данные
        for i, data in enumerate(block_data):
            row = 6 + i
            if i > 0:
                worksheet.insert_rows(row, 1)
            
            worksheet[f"A{row}"] = i + 1
            worksheet[f"B{row}"] = data['block_name']
            worksheet[f"C{row}"] = f"{data['percentage']}%"
            worksheet[f"D{row}"] = data['total_employees']
            worksheet[f"E{row}"] = data['completed_employees']
            worksheet[f"F{row}"] = data['remaining_employees']
            worksheet[f"G{row}"] = data['update_date']
        
        # Границы
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                           top=Side(style='thin'), bottom=Side(style='thin'))
        for i in range(len(block_data)):
            row = 6 + i
            for col in range(1, 8):
                worksheet.cell(row=row, column=col).border = thin_border
        
        # Формулы итого
        if block_data:
            total_row = 6 + len(block_data) + 1
            data_end_row = 6 + len(block_data) - 1
            worksheet[f"C{total_row}"] = f'=ROUND(E{total_row}/D{total_row}*100,0)&"%"'
            worksheet[f"D{total_row}"] = f'=SUM(D6:D{data_end_row})'
            worksheet[f"E{total_row}"] = f'=SUM(E6:E{data_end_row})'
            worksheet[f"F{total_row}"] = f'=SUM(F6:F{data_end_row})'
        
        workbook.save(output_path)
        workbook.close()
        return True

    def create_general_report_from_blocks_with_rules(self, block_data: List[Dict], output_path: str) -> bool:
        """Создает общий отчет используя rules"""
        template_path = Path(self.config.general_report_template)
        if not template_path.exists():
            raise FileNotFoundError(f"Шаблон общего отчета не найден: {template_path}")
        
        self.directory_manager.ensure_directory_exists(Path(output_path).parent)
        shutil.copy2(template_path, output_path)
        
        # Загружаем rules общего отчета
        rules = self._load_filling_rules(str(template_path))
        
        workbook = openpyxl.load_workbook(output_path)
        
        # Используем DataMapper для динамического маппинга заголовка
        general_data = self.data_mapper.map_general_header_data(block_data)
        
        self._apply_rules_to_template(workbook, rules, general_data)
        
        # Заполняем таблицу данных (header правила)
        self._fill_general_report_table_dynamic(workbook, block_data, rules)
        
        workbook.save(output_path)
        workbook.close()
        return True

    def _copy_row_formatting(self, worksheet, source_row: int, target_row: int, max_col: int = 20):
        """Копирует форматирование из исходной строки в целевую"""
        try:
            for col in range(1, max_col + 1):
                source_cell = worksheet.cell(row=source_row, column=col)
                target_cell = worksheet.cell(row=target_row, column=col)
                
                # Копируем все стили с проверкой на None
                if source_cell.has_style:
                    if source_cell.font is not None:
                        target_cell.font = source_cell.font
                    if source_cell.border is not None:
                        target_cell.border = source_cell.border
                    if source_cell.fill is not None:
                        target_cell.fill = source_cell.fill
                    if source_cell.number_format is not None:
                        target_cell.number_format = source_cell.number_format
                    if source_cell.protection is not None:
                        target_cell.protection = source_cell.protection
                    if source_cell.alignment is not None:
                        target_cell.alignment = source_cell.alignment
        except Exception as e:
            self.logger.warning(f"Ошибка копирования форматирования строки {source_row} в {target_row}: {e}")

    def _fill_general_report_table_dynamic(self, workbook, block_data: List[Dict], rules: Dict[str, Dict[str, str]]):
        """Динамически заполняет таблицу общего отчета используя DataMapper с сохранением форматирования"""
        if 'Report' not in workbook.sheetnames:
            return
        
        worksheet = workbook['Report']
        header_rules = rules.get('header', {})
        
        # Собираем mapping: имя поля -> столбец и строка
        column_mapping = {}
        for cell_address, field_name in header_rules.items():
            try:
                is_formula, clean_address, sheet_name = self._parse_cell_address(cell_address)
                if sheet_name is None:
                    sheet_name = 'Report'
                
                if sheet_name == 'Report':
                    col_match = re.search(r'([A-Z]+)', clean_address)
                    row_match = re.search(r'(\d+)', clean_address)
                    if col_match and row_match:
                        column_mapping[field_name] = (col_match.group(1), int(row_match.group(1)))
            except Exception as e:
                self.logger.warning(f"Ошибка парсинга адреса {cell_address}: {e}")
        
        if not column_mapping:
            return
        
        # Определяем строку-шаблон для копирования форматирования (обычно это строка заголовков)
        template_row = 8  # Предполагаем, что заголовки в 8-й строке
        start_data_row = 9  # Данные начинаются с 9-й строки
        
        # Находим максимальный столбец для копирования форматирования
        max_col = 1
        for col_str, _ in column_mapping.values():
            col_num = 0
            for char in col_str:
                col_num = col_num * 26 + (ord(char.upper()) - ord('A') + 1)
            max_col = max(max_col, col_num)
        
        # Заполняем данные используя DataMapper
        for i, data in enumerate(block_data):
            row = start_data_row + i  # Начинаем с 9-й строки (после заголовков)
            
            # Вставляем строку если нужно
            if i > 0:
                worksheet.insert_rows(row, 1)
                # Копируем форматирование из строки-шаблона
                self._copy_row_formatting(worksheet, template_row, row, max_col)
            
            # Получаем данные через DataMapper
            row_data = self.data_mapper.map_block_data_to_rules(data, i, '')
            
            # Заполняем каждое поле
            for field_name, (col, header_row) in column_mapping.items():
                cell_address = f"{col}{row}"
                value = row_data.get(field_name, '')
                
                # Преобразуем значение к правильному типу
                converted_value = self._convert_value_type(value)
                worksheet[cell_address] = converted_value
        
        # Применяем границы (если нужно дополнительно)
        if block_data:
            self._apply_borders_to_general_table(worksheet, len(block_data))
    
    def _fill_general_report_table(self, workbook, block_data: List[Dict], rules: Dict[str, Dict[str, str]]):
        """Заполняет таблицу общего отчета по rules с сохранением форматирования"""
        if 'Report' not in workbook.sheetnames:
            return
        
        worksheet = workbook['Report']
        header_rules = rules.get('header', {})
        
        # Собираем mapping: имя поля -> столбец и строка
        column_mapping = {}
        for cell_address, field_name in header_rules.items():
            try:
                is_formula, clean_address, sheet_name = self._parse_cell_address(cell_address)
                if sheet_name is None:
                    sheet_name = 'Report'
                
                if sheet_name == 'Report':
                    col_match = re.search(r'([A-Z]+)', clean_address)
                    row_match = re.search(r'(\d+)', clean_address)
                    if col_match and row_match:
                        column_mapping[field_name] = (col_match.group(1), int(row_match.group(1)))
            except Exception as e:
                self.logger.warning(f"Ошибка парсинга адреса {cell_address}: {e}")
        
        if not column_mapping:
            return
        
        # Определяем строку-шаблон для копирования форматирования
        template_row = 8  # Предполагаем, что заголовки в 8-й строке
        start_data_row = 9  # Данные начинаются с 9-й строки
        
        # Находим максимальный столбец для копирования форматирования
        max_col = 1
        for col_str, _ in column_mapping.values():
            col_num = 0
            for char in col_str:
                col_num = col_num * 26 + (ord(char.upper()) - ord('A') + 1)
            max_col = max(max_col, col_num)
        
        # Заполняем данные
        for i, data in enumerate(block_data):
            row = start_data_row + i  # Начинаем с 9-й строки (после заголовков)
            
            # Вставляем строку если нужно
            if i > 0:
                worksheet.insert_rows(row, 1)
                # Копируем форматирование из строки-шаблона
                self._copy_row_formatting(worksheet, template_row, row, max_col)
            
            # Заполняем каждое поле
            for field_name, (col, header_row) in column_mapping.items():
                cell_address = f"{col}{row}"
                
                if field_name == 'row_number2':
                    value = i + 1
                elif field_name == 'report_department1':
                    value = data.get('block_name', '')
                elif field_name == 'percentage':
                    value = f"{data.get('percentage', 0)}%"
                elif field_name == 'employees_count':
                    value = data.get('total_employees', 0)
                elif field_name == 'correct_filled':
                    value = data.get('completed_employees', 0)
                elif field_name == 'incorrect_filled':
                    value = data.get('employees_incorrect', 0)
                elif field_name == 'not_filled':
                    value = data.get('employees_not_filled', 0)
                elif field_name == 'update_date':
                    value = data.get('update_date', '')
                else:
                    value = ''
                
                # Преобразуем значение к правильному типу
                converted_value = self._convert_value_type(value)
                worksheet[cell_address] = converted_value
        
        # Применяем границы (если нужно дополнительно)
        if block_data:
            self._apply_borders_to_general_table(worksheet, len(block_data))

    def _apply_borders_to_general_table(self, worksheet, data_count: int):
        """Применяет границы к таблице общего отчета"""
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                           top=Side(style='thin'), bottom=Side(style='thin'))
        
        start_row = 9
        end_row = start_row + data_count - 1
        
        # Определяем количество столбцов по правилам
        max_col = 8  # Фиксированное количество столбцов для общего отчета
        
        # Применяем границы
        for row in range(start_row, end_row + 1):
            for col in range(1, max_col + 1):
                worksheet.cell(row=row, column=col).border = thin_border

    def _get_cell_value(self, worksheet, cell_address: str):
        """Безопасно получает значение ячейки"""
        try:
            return worksheet[cell_address].value
        except Exception:
            return None

    def _parse_date(self, date_value) -> Optional[date]:
        """Парсит дату из различных форматов"""
        if not date_value:
            return None
        
        if isinstance(date_value, date):
            return date_value
        if isinstance(date_value, datetime):
            return date_value.date()
        
        date_str = str(date_value).strip()
        if not date_str:
            return None
        
        formats = ["%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"]
        
        for fmt in formats:
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue
        
        return None

    def generate_output_filename(self, employee: Dict[str, str]) -> str:
        """Генерирует имя файла для сотрудника"""
        clean_fio = self._clean_filename(employee['ФИО работника'])
        clean_tab_num = self._clean_filename(employee['Табельный номер'])
        return f"{clean_fio} ({clean_tab_num}).xlsx"

    def generate_block_report_filename(self, block_name: str) -> str:
        """Генерирует имя файла отчета по блоку"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        clean_block_name = self._clean_filename(block_name)
        return f"Отчет по блоку_{clean_block_name}_{timestamp}.xlsx"

    def _clean_filename(self, filename: str) -> str:
        """Очищает имя файла от недопустимых символов"""
        if not filename:
            return "unnamed"
        
        clean_name = re.sub(r'[\\/:*?"<>|]', '_', filename).strip()
        return clean_name[:100] if len(clean_name) > 100 else clean_name or "unnamed"
