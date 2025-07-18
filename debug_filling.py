#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Детальная отладка заполнения файлов
"""

import sys
import os
sys.path.append('.')

from core.excel_handler import ExcelHandler
from core.processor import VacationProcessor
import openpyxl

def debug_file_creation():
    """Отладка создания одного файла"""
    
    # Создаем тестовые данные
    test_employee = {
        'Табельный номер': '12345',
        'ФИО работника': 'Тестовый Сотрудник',
        'Должность': 'Тестовая Должность',
        'Подразделение 1': 'Тестовый Отдел',
        'Подразделение 2': '',
        'Подразделение 3': '',
        'Подразделение 4': ''
    }
    
    # Создаем временную папку
    output_dir = "temp_debug"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    try:
        # Создаем файл
        file_path = os.path.join(output_dir, f"{test_employee['ФИО работника']} ({test_employee['Табельный номер']}).xlsx")
        
        print(f"Создаем файл: {file_path}")
        
        # Создаем файл через ExcelHandler
        from config import Config
        config = Config()
        config.load()
        
        excel_handler = ExcelHandler(config)
        success = excel_handler.create_employee_file(
            employee=test_employee,
            output_path=file_path
        )
        
        print(f"Результат создания: {success}")
        
        if success and os.path.exists(file_path):
            print(f"Файл создан: {file_path}")
            
            # Проверяем содержимое
            wb = openpyxl.load_workbook(file_path)
            ws = wb['Расчёт']
            
            print("\nПроверяем заполнение:")
            print(f"B4 (Подразделение 1): '{ws['B4'].value}'")
            print(f"B10 (ФИО): '{ws['B10'].value}'")
            print(f"C4 (формула): '{ws['C4'].value}'")
            print(f"C10 (формула): '{ws['C10'].value}'")
            
            # Проверяем другие ячейки
            print(f"B5 (Подразделение 2): '{ws['B5'].value}'")
            print(f"B6 (Подразделение 3): '{ws['B6'].value}'")
            print(f"B8 (Должность): '{ws['B8'].value}'")
            print(f"B9 (Табельный номер): '{ws['B9'].value}'")
            
            wb.close()
        else:
            print("Файл не был создан или создание не удалось")
            
    except Exception as e:
        print(f"Ошибка: {e}")
        import traceback
        traceback.print_exc()
    
    finally:
        # Удаляем временную папку
        if os.path.exists(output_dir):
            import shutil
            shutil.rmtree(output_dir)

if __name__ == "__main__":
    debug_file_creation() 