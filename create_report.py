#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Создание отчета по блоку - Автономный скрипт
Использует файлы сотрудников в текущей папке для создания отчета по подразделению
"""

import os
import sys
import shutil
import re
from pathlib import Path
from datetime import datetime, date, timedelta
from typing import List, Optional, Dict, Any

# Проверка зависимостей
try:
    import openpyxl
    from openpyxl.styles import Border, Side
except ImportError:
    print("ОШИБКА: Не установлена библиотека openpyxl")
    print("Установите: pip install openpyxl")
    import time
    time.sleep(3)
    sys.exit(1)

# =====================================================
# КОНСТАНТЫ
# =====================================================

TEMPLATE_PATH = r"M:\Подразделения\АУП\Стажерская программа\Отпуск Р7\templates\block_report_template v3.xlsx"
TARGET_YEAR = 2026
DAYS_IN_MONTHS = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
MONTH_NAMES = ['Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
               'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь']

VALIDATION_STATUSES = {
    "not_filled": "Форма не заполнена",
    "filled_incorrect": "Форма заполнена некорректно", 
    "filled_correct": "Форма заполнена корректно"
}

CALENDAR_START_COL = 12
CALENDAR_MONTH_ROW = 7
CALENDAR_DAY_ROW = 8
EMPLOYEE_DATA_START_ROW = 9

# =====================================================
# МОДЕЛИ
# =====================================================

class VacationPeriod:
    def __init__(self, start_date: date, end_date: date, days: int = 0):
        self.start_date = start_date
        self.end_date = end_date
        self.days = days if days > 0 else (end_date - start_date).days + 1

class VacationInfo:
    def __init__(self, employee: Dict[str, str], periods: List[VacationPeriod] = None):
        self.employee = employee
        self.periods = periods or []
        self.status = "Форма не заполнена"
        self.validation_errors = []
    
    @property
    def total_days(self) -> int:
        return sum(period.days for period in self.periods)
    
    @property
    def periods_count(self) -> int:
        return len(self.periods)

# =====================================================
# ФУНКЦИИ
# =====================================================

def load_rules(template_path: str) -> Dict[str, Dict[str, str]]:
    """Загружает правила заполнения из листа rules"""
    rules = {'value': {}, 'header': {}}
    
    try:
        workbook = openpyxl.load_workbook(template_path, data_only=False)
        
        if 'rules' not in workbook.sheetnames:
            print(f"ПРЕДУПРЕЖДЕНИЕ: Лист rules не найден в {template_path}")
            workbook.close()
            return rules
        
        rules_sheet = workbook['rules']
        
        for row in range(2, rules_sheet.max_row + 1):
            target_address = rules_sheet.cell(row=row, column=1).value
            source_field = rules_sheet.cell(row=row, column=2).value
            rule_type = rules_sheet.cell(row=row, column=3).value
            
            if target_address and source_field and rule_type:
                target_address = str(target_address).strip()
                source_field = str(source_field).strip()
                rule_type = str(rule_type).strip().lower()
                
                if rule_type in ['value', 'header']:
                    rules[rule_type][target_address] = source_field
        
        workbook.close()
        return rules
        
    except Exception as e:
        print(f"ОШИБКА: Не удалось загрузить rules из {template_path}: {e}")
        return rules

def parse_cell_address(address: str) -> tuple:
    """Парсит адрес ячейки"""
    is_formula = address.startswith('=')
    
    if is_formula:
        formula = address[1:]
        if '!' in formula:
            sheet_part, cell_part = formula.split('!', 1)
            sheet_name = sheet_part.strip("'\"")
            clean_address = cell_part.strip()
        else:
            sheet_name = None
            clean_address = formula.strip()
    else:
        sheet_name = None
        clean_address = address.strip()
    
    return (is_formula, clean_address, sheet_name)

def convert_value_type(value: Any) -> Any:
    """Преобразует значение к правильному типу для Excel"""
    if value is None or value == '':
        return ''
    
    if isinstance(value, (int, float)):
        return float(value)
    
    str_value = str(value).strip()
    if not str_value:
        return ''
    
    try:
        clean_value = str_value.replace(' ', '').replace('\xa0', '').replace(',', '.')
        return float(clean_value)
    except (ValueError, TypeError):
        return str_value

def fill_cell(workbook, sheet_name: str, address: str, value: Any):
    """Заполняет ячейку значением"""
    converted_value = convert_value_type(value)
    
    try:
        if sheet_name and sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.worksheets[0]
        
        if re.match(r'^[A-Z]+[0-9]+$', address):
            cell = worksheet[address]
            cell.value = converted_value
            if isinstance(converted_value, (int, float)):
                cell.data_type = 'n'
            elif isinstance(converted_value, str) and converted_value != '':
                cell.data_type = 's'
        elif ':' in address:
            start_cell = address.split(':')[0]
            if re.match(r'^[A-Z]+[0-9]+$', start_cell):
                cell = worksheet[start_cell]
                cell.value = converted_value
        else:
            cell = worksheet[address]
            cell.value = converted_value
            
    except Exception as e:
        print(f"ПРЕДУПРЕЖДЕНИЕ: Ошибка при заполнении {address}: {e}")

def apply_rules(workbook, rules: Dict[str, Dict[str, str]], data_dict: Dict[str, Any]):
    """Применяет правила заполнения"""
    for rule_type, rule_items in rules.items():
        if rule_type == 'value':
            for cell_address, field_name in rule_items.items():
                value = data_dict.get(field_name, '')
                try:
                    is_formula, clean_address, sheet_name = parse_cell_address(cell_address)
                    fill_cell(workbook, sheet_name, clean_address, value)
                except Exception as e:
                    print(f"ПРЕДУПРЕЖДЕНИЕ: Ошибка при заполнении {cell_address}: {e}")

def is_employee_file(filename: str) -> bool:
    """Проверяет маску файла сотрудника"""
    pattern = r"^.+\s\(\d+\)\.xlsx$"
    return bool(re.match(pattern, filename, re.IGNORECASE))

def scan_employee_files(directory: str) -> List[str]:
    """Сканирует файлы сотрудников"""
    employee_files = []
    directory_path = Path(directory)
    
    print(f"Сканирование папки: {directory_path.absolute()}")
    
    for file_path in directory_path.iterdir():
        if (file_path.is_file() and 
            file_path.suffix.lower() == '.xlsx' and 
            not file_path.name.startswith('~$') and 
            is_employee_file(file_path.name)):
            employee_files.append(str(file_path))
    
    print(f"Найдено файлов сотрудников: {len(employee_files)}")
    return employee_files

def parse_date(date_value) -> Optional[date]:
    """Парсит дату"""
    if not date_value:
        return None
    
    if isinstance(date_value, date):
        return date_value
    if isinstance(date_value, datetime):
        return date_value.date()
    
    date_str = str(date_value).strip()
    if not date_str:
        return None
    
    formats = ["%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"]
    
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    
    return None

def get_cell_value(worksheet, cell_address: str):
    """Получает значение ячейки"""
    try:
        return worksheet[cell_address].value
    except Exception:
        return None

def read_vacation_info(file_path: str, employee_rules: Dict = None) -> Optional[VacationInfo]:
    """Читает информацию об отпусках из файла сотрудника"""
    try:
        # Загружаем rules из файла если не переданы
        if employee_rules is None:
            employee_rules = load_rules(file_path)
        
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        worksheet = workbook.worksheets[0]
        
        # Читаем данные сотрудника по rules
        employee = {}
        value_rules = employee_rules.get('value', {})
        
        for cell_address, field_name in value_rules.items():
            try:
                is_formula, clean_address, sheet_name = parse_cell_address(cell_address)
                if sheet_name and sheet_name in workbook.sheetnames:
                    ws = workbook[sheet_name]
                else:
                    ws = worksheet
                
                value = get_cell_value(ws, clean_address)
                employee[field_name] = str(value).strip() if value is not None else ""
            except Exception:
                employee[field_name] = ""
        
        # Читаем периоды отпусков из строк 15-29
        periods = []
        for row in range(15, 30):
            start_date_value = get_cell_value(worksheet, f"C{row}")
            end_date_value = get_cell_value(worksheet, f"D{row}")
            
            if not start_date_value or not end_date_value:
                continue
            
            try:
                start_date = parse_date(start_date_value)
                end_date = parse_date(end_date_value)
                
                if start_date and end_date:
                    days = (end_date - start_date).days + 1
                    periods.append(VacationPeriod(start_date, end_date, days))
            except Exception:
                continue
        
        # Читаем статус из B12
        status_value = get_cell_value(worksheet, "B12")
        status_text = str(status_value).strip() if status_value else ""
        
        vacation_info = VacationInfo(employee=employee, periods=periods)
        
        # Определяем статус
        if status_text == VALIDATION_STATUSES["not_filled"]:
            vacation_info.status = "Форма не заполнена"
        elif status_text == VALIDATION_STATUSES["filled_correct"]:
            vacation_info.status = "Форма заполнена корректно"
        elif status_text == VALIDATION_STATUSES["filled_incorrect"]:
            vacation_info.status = "Форма заполнена некорректно"
        else:
            if not periods:
                vacation_info.status = "Форма не заполнена"
            else:
                vacation_info.status = "Форма заполнена некорректно"
        
        if status_text and vacation_info.status != "Форма заполнена корректно":
            vacation_info.validation_errors = [status_text]
        
        workbook.close()
        return vacation_info
        
    except Exception as e:
        print(f"ОШИБКА: Не удалось прочитать файл {file_path}: {e}")
        return None

def get_calendar_column(target_date: date, start_col: int) -> Optional[int]:
    """Вычисляет столбец для даты в календаре"""
    if target_date.year != TARGET_YEAR:
        return None
    
    col_offset = sum(DAYS_IN_MONTHS[:target_date.month - 1]) + target_date.day - 1
    return start_col + col_offset

def fill_calendar_matrix(worksheet, vacation_infos: List[VacationInfo]):
    """Заполняет календарную матрицу"""
    try:
        # Заголовки месяцев и дней
        col_offset = 0
        for month_idx, month_name in enumerate(MONTH_NAMES):
            month_col = CALENDAR_START_COL + col_offset
            worksheet.cell(row=CALENDAR_MONTH_ROW, column=month_col, value=month_name)
            
            days_in_month = DAYS_IN_MONTHS[month_idx]
            for day in range(1, days_in_month + 1):
                day_col = CALENDAR_START_COL + col_offset + day - 1
                worksheet.cell(row=CALENDAR_DAY_ROW, column=day_col, value=day)
            
            col_offset += days_in_month
        
        # Отпуска сотрудников
        for emp_idx, vacation_info in enumerate(vacation_infos):
            emp_row = EMPLOYEE_DATA_START_ROW + emp_idx
            for period in vacation_info.periods:
                current_date = period.start_date
                while current_date <= period.end_date:
                    if current_date.year == TARGET_YEAR:
                        day_col = get_calendar_column(current_date, CALENDAR_START_COL)
                        if day_col:
                            worksheet.cell(row=emp_row, column=day_col, value=1)
                    
                    current_date += timedelta(days=1)
                    if current_date > period.end_date:
                        break
                        
    except Exception as e:
        print(f"ПРЕДУПРЕЖДЕНИЕ: Ошибка заполнения календаря: {e}")

def fill_table_by_prefix(worksheet, data_list: List, rules: Dict, prefix: str, row_data_func):
    """Заполняет таблицу по префиксу"""
    header_rules = rules.get('header', {})
    
    # Собираем mapping полей
    column_mapping = {}
    for cell_address, field_name in header_rules.items():
        if field_name.startswith(prefix):
            if cell_address.startswith('=') and '!' in cell_address:
                sheet_and_cell = cell_address[1:]
                if '!' in sheet_and_cell:
                    sheet_part, cell_part = sheet_and_cell.split('!', 1)
                    col_match = re.search(r'([A-Z]+)', cell_part)
                    row_match = re.search(r'(\d+)', cell_part)
                    if col_match and row_match:
                        column_mapping[field_name] = (col_match.group(1), int(row_match.group(1)))
            else:
                col_match = re.search(r'([A-Z]+)', cell_address)
                row_match = re.search(r'(\d+)', cell_address)
                if col_match and row_match:
                    column_mapping[field_name] = (col_match.group(1), int(row_match.group(1)))
    
    # Заполняем данные
    for i, data_item in enumerate(data_list):
        row_data = row_data_func(data_item, i, prefix)
        
        for key, value in row_data.items():
            if key in column_mapping:
                col, header_row = column_mapping[key]
                data_row = header_row + 1 + i
                cell_address = f"{col}{data_row}"
                
                converted_value = convert_value_type(value)
                worksheet[cell_address] = converted_value

def get_row_data(item, index: int, prefix: str) -> Dict[str, Any]:
    """Получает данные строки для заполнения"""
    if hasattr(item, 'employee'):  # VacationInfo
        employee = item.employee
        return {
            f'{prefix}employee_name': employee.get('ФИО работника', ''),
            f'{prefix}tab_number': employee.get('Табельный номер', ''),
            f'{prefix}position': employee.get('Должность', ''),
            f'{prefix}department1': employee.get('Подразделение 1', ''),
            f'{prefix}department2': employee.get('Подразделение 2', ''),
            f'{prefix}department3': employee.get('Подразделение 3', ''),
            f'{prefix}department4': employee.get('Подразделение 4', ''),
            f'{prefix}status': item.status,
            f'{prefix}total_days': item.total_days,
            f'{prefix}periods_count': item.periods_count,
            f'{prefix}row_number': index + 1
        }
    else:  # Normalized data для Print
        employee = item.get('employee', {})
        start_date = item.get('start_date')
        end_date = item.get('end_date')
        days = item.get('days', 0)
        
        return {
            f'{prefix}employee_name': employee.get('ФИО работника', ''),
            f'{prefix}tab_number': employee.get('Табельный номер', ''),
            f'{prefix}position': employee.get('Должность', ''),
            f'{prefix}start_date': start_date.strftime('%d.%m.%Y') if start_date else '',
            f'{prefix}end_date': end_date.strftime('%d.%m.%Y') if end_date else '',
            f'{prefix}duration': str(days) if days else '',
            f'{prefix}signature': '',
            f'{prefix}acknowledgment_date': '',
            f'{prefix}notes': '',
            f'{prefix}row_number': index + 1
        }

def normalize_vacation_data(vacation_infos: List[VacationInfo]) -> List[Dict]:
    """Нормализует данные - каждый период = строка"""
    normalized_data = []
    for vacation_info in vacation_infos:
        emp = vacation_info.employee
        if not vacation_info.periods:
            normalized_data.append({
                'employee': emp, 'period_num': 0, 'start_date': None, 'end_date': None, 'days': 0
            })
        else:
            for period_idx, period in enumerate(vacation_info.periods, 1):
                normalized_data.append({
                    'employee': emp, 'period_num': period_idx,
                    'start_date': period.start_date, 'end_date': period.end_date, 'days': period.days
                })
    return normalized_data

def create_block_report(block_name: str, vacation_infos: List[VacationInfo], output_path: str) -> bool:
    """Создает отчет по блоку"""
    try:
        # Копируем шаблон
        shutil.copy2(TEMPLATE_PATH, output_path)
        
        # Загружаем rules
        rules = load_rules(TEMPLATE_PATH)
        
        # Открываем файл
        workbook = openpyxl.load_workbook(output_path)
        
        # Заполняем заголовок
        total_employees = len(vacation_infos)
        employees_filled = sum(1 for info in vacation_infos if info.status != "Форма не заполнена")
        employees_correct = sum(1 for info in vacation_infos if info.status == "Форма заполнена корректно")
        
        report_data = {
            'block_name': block_name,
            'update_date': datetime.now().strftime('%d.%m.%Y %H:%M'),
            'total_employees': str(total_employees),
            'employees_filled': str(employees_filled),
            'employees_correct': str(employees_correct)
        }
        
        apply_rules(workbook, rules, report_data)
        
        # Заполняем таблицы
        if 'Report' in workbook.sheetnames:
            fill_table_by_prefix(workbook['Report'], vacation_infos, rules, 'report_', get_row_data)
            fill_calendar_matrix(workbook['Report'], vacation_infos)
        
        if 'Print' in workbook.sheetnames:
            normalized_data = normalize_vacation_data(vacation_infos)
            fill_table_by_prefix(workbook['Print'], normalized_data, rules, 'print_', get_row_data)
        
        workbook.save(output_path)
        workbook.close()
        
        print(f"Отчет создан: {output_path}")
        return True
        
    except Exception as e:
        print(f"ОШИБКА: Ошибка создания отчета: {e}")
        return False

# =====================================================
# ГЛАВНАЯ ФУНКЦИЯ
# =====================================================

def main():
    """Главная функция"""
    print("=" * 60)
    print("  СОЗДАНИЕ ОТЧЕТА ПО БЛОКУ")
    print("=" * 60)
    print()
    
    # 1. Проверяем шаблон
    print("1. Проверка шаблона...")
    if not Path(TEMPLATE_PATH).exists():
        print(f"ОШИБКА: Шаблон не найден: {TEMPLATE_PATH}")
        import time
        time.sleep(3)
        return
    print("Шаблон найден")
    
    # 2. Сканируем файлы
    current_dir = os.getcwd()
    print(f"2. Текущая папка: {current_dir}")
    print("3. Поиск файлов сотрудников...")
    
    employee_files = scan_employee_files(current_dir)
    if not employee_files:
        print("ОШИБКА: Не найдено файлов сотрудников по маске 'Имя (цифры).xlsx'")
        import time
        time.sleep(3)
        return
    
    # 3. Читаем данные
    print("4. Чтение данных...")
    vacation_infos = []
    employee_rules = None
    
    for i, file_path in enumerate(employee_files, 1):
        print(f"   Обработка {i}/{len(employee_files)}: {Path(file_path).name}")
        
        # Загружаем rules из первого файла
        if employee_rules is None:
            employee_rules = load_rules(file_path)
            print("Загружены rules из файла сотрудника")
        
        vacation_info = read_vacation_info(file_path, employee_rules)
        if vacation_info and vacation_info.employee.get('ФИО работника'):
            vacation_infos.append(vacation_info)
    
    if not vacation_infos:
        print("ОШИБКА: Не удалось прочитать файлы сотрудников")
        import time
        time.sleep(3)
        return
    
    print(f"Успешно обработано: {len(vacation_infos)} файлов")
    
    # 4. Определяем название блока
    block_name = "Неизвестное подразделение"
    for vacation_info in vacation_infos:
        dept1 = vacation_info.employee.get('Подразделение 1', '').strip()
        if dept1:
            block_name = dept1
            break
    
    if block_name == "Неизвестное подразделение":
        current_folder_name = Path(current_dir).name
        if current_folder_name and current_folder_name != '.':
            block_name = current_folder_name
    
    print(f"5. Название блока: {block_name}")
    
    # 5. Создаем отчет
    print("6. Создание отчета...")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"Отчет по блоку_{block_name}_{timestamp}.xlsx"
    
    # Очищаем имя файла
    invalid_chars = r'[\\/:*?"<>|]'
    output_filename = re.sub(invalid_chars, '_', output_filename)
    output_path = Path(current_dir) / output_filename
    
    success = create_block_report(block_name, vacation_infos, str(output_path))
    
    if success:
        print()
        print("=" * 60)
        print("  ОТЧЕТ СОЗДАН!")
        print("=" * 60)
        print(f"Файл: {output_filename}")
        print(f"Подразделение: {block_name}")
        print(f"Сотрудников: {len(vacation_infos)}")
        print(f"Целевой год: {TARGET_YEAR}")
        
        # Статистика
        status_counts = {}
        for vi in vacation_infos:
            status = vi.status
            status_counts[status] = status_counts.get(status, 0) + 1
        
        print("Статистика:")
        for status, count in status_counts.items():
            print(f"  {status}: {count}")
        
        print("\nОтчет создан в текущей папке.")
    else:
        print("ОШИБКА: Не удалось создать отчет")
    
    print()
    import time
    time.sleep(2)  # Пауза 2 сек чтобы увидеть результат

if __name__ == "__main__":
    main()